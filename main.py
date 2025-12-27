#!/usr/bin/env python3
"""
Swiggy App Store Review Trend Analysis System
Analyzes reviews from Google Play Store using Agentic AI approaches
"""

import json
import os
import sys
import re
import argparse
import time
import asyncio
import multiprocessing
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any
from concurrent.futures import ThreadPoolExecutor, as_completed

# Performance monitoring
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
    print("Warning: psutil not available, using default concurrency settings")

# Third-party imports
import pandas as pd
from google_play_scraper import app, reviews, Sort
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Import and apply hardware profile optimization
from config.hardware_profiles import apply_profile
hardware_profile = apply_profile()  # Auto-detect and configure

# Import LLM abstraction
from config.llm_client import get_llm_client, OllamaClient, AsyncOllamaClient

# Initialize LLM client (Ollama or fallback)
client = get_llm_client()


def get_optimal_concurrency() -> dict:
    """
    Auto-detect optimal concurrency settings based on hardware

    Returns:
        dict with 'cpu_count', 'available_memory_gb', 'max_concurrent', 'batch_size'
    """
    cpu_count = multiprocessing.cpu_count()

    # Get available memory
    if PSUTIL_AVAILABLE:
        available_memory_gb = psutil.virtual_memory().available / (1024**3)
    else:
        available_memory_gb = 8.0  # Conservative default

    # Heuristics based on hardware capabilities
    # Each worker needs ~500MB RAM for Ollama inference
    max_workers_by_cpu = cpu_count * 2  # Hyperthreading benefit for async I/O
    max_workers_by_memory = int(available_memory_gb / 0.5)

    # Conservative: Take minimum to avoid OOM
    recommended_workers = min(max_workers_by_cpu, max_workers_by_memory, 32)

    # For async I/O, we can go higher than thread-based parallelism
    recommended_concurrent = recommended_workers * 2

    # Batch size: Smaller batches with more workers for better parallelism
    if recommended_workers <= 8:
        batch_size = 20
    elif recommended_workers <= 16:
        batch_size = 15
    else:
        batch_size = 10

    # Allow environment variable overrides
    config = {
        'cpu_count': cpu_count,
        'available_memory_gb': round(available_memory_gb, 1),
        'max_workers': int(os.getenv('MAX_WORKERS', recommended_workers)),
        'max_concurrent': int(os.getenv('MAX_CONCURRENT', recommended_concurrent)),
        'batch_size': int(os.getenv('BATCH_SIZE', batch_size))
    }

    return config


# Configuration
SWIGGY_APP_ID = "in.swiggy.android"
DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")
CACHE_DIR = Path("cache")
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
CACHE_DIR.mkdir(exist_ok=True)

# Prompts
EXTRACTION_PROMPT = """Extract ALL topics from this Swiggy review. Prioritize HIGH RECALL.

Include context in topic names (e.g., "delivery partner rude" not just "rude").
Detect sarcasm: "Great job delivering cold food" → "food delivered cold"
Return max 5 most important topics.

Review: {review_text}

Output JSON array ONLY (no markdown, just raw JSON):
[{{"topic": "descriptive phrase", "category": "issue|request|feedback"}}]"""

BATCH_EXTRACTION_PROMPT = """Extract ALL topics from these app reviews. Prioritize HIGH RECALL.

Include context in topic names (e.g., "delivery partner rude" not just "rude").
Detect sarcasm: "Great job delivering cold food" → "food delivered cold"
Return max 5 most important topics per review.

Reviews:
{reviews_batch}

Output JSON object ONLY (no markdown, just raw JSON):
{{"reviews": [
  {{"review_id": "0", "topics": [{{"topic": "descriptive phrase", "category": "issue|request|feedback"}}]}},
  {{"review_id": "1", "topics": [{{"topic": "descriptive phrase", "category": "issue|request|feedback"}}]}},
  ...
]}}"""

CONSOLIDATION_PROMPT = """You are consolidating topics from app reviews. BE EXTREMELY AGGRESSIVE - aim for 15-25 final topics MAXIMUM.

Here are {num_topics} extracted topics:
{topics_list}

CRITICAL RULES - MERGE EVERYTHING SIMILAR:
1. **ALL positive feedback** → "Positive feedback" (good, great, excellent, amazing, love, helpful, friendly, fast, etc.)
2. **ALL negative delivery partner behavior** → "Delivery partner unprofessional" (rude, impolite, disrespectful, unprofessional, etc.)
3. **ALL delivery delays** → "Delivery delay" (late, delayed, slow, 2 hours, extreme delay, etc.)
4. **ALL food quality issues** → Merge into 2-3 categories ONLY:
   - "Food temperature issues" (cold, hot, lukewarm)
   - "Food freshness issues" (stale, spoiled, rotten, old)
   - "Food quality issues" (bad quality, taste, portion)
5. **ALL app technical issues** → Merge into 2 categories:
   - "App crashes/freezes" (crash, freeze, not responding, stuck)
   - "App performance issues" (slow, laggy, buggy, glitches)
6. **ALL feature removal/requests** → Use ONE topic per feature:
   - "10 minute delivery removed"
   - "24/7 service request"
7. **Ignore ALL grammar differences**: tense, plural, word order, articles
8. **Merge similar sentiment**: "good", "great", "excellent", "awesome" → "Positive feedback"

AGGRESSIVE MERGING EXAMPLES:
- "good service", "great app", "excellent delivery", "love it", "awesome", "amazing" → "Positive feedback"
- "delivery guy rude", "rude rider", "impolite partner", "disrespectful delivery" → "Delivery partner unprofessional"
- "food cold", "cold food", "food not hot", "lukewarm food" → "Food temperature issues"
- "app crash", "app freezes", "app not working", "app stuck" → "App crashes/freezes"
- "app slow", "app laggy", "app buggy", "app has bugs", "app glitchy" → "App performance issues"

TARGET: 15-25 CANONICAL TOPICS MAXIMUM. Be RUTHLESS in merging!

OUTPUT (JSON only, no markdown):
{{
  "canonical_topics": [
    {{
      "canonical_name": "Delivery delay",
      "variations": ["delivery delay 2 hours", "delivery delayed 2 hours", "2 hour delivery wait", "delivery extremely delayed"]
    }},
    ...
  ]
}}"""


def get_cache_file_path(app_id: str) -> Path:
    """Get cache file path for an app. Creates per-app subdirectory."""
    app_cache_dir = CACHE_DIR / app_id
    app_cache_dir.mkdir(exist_ok=True)
    return app_cache_dir / "reviews_cache.json"


def load_cached_reviews(app_id: str) -> tuple[list, datetime]:
    """
    Load cached reviews for an app.
    Returns: (reviews_list, last_update_time)
    """
    cache_file = get_cache_file_path(app_id)

    if cache_file.exists():
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                reviews = data.get('reviews', [])
                last_update = datetime.fromisoformat(data.get('last_update', datetime.now().isoformat()))

                # Convert all datetime string fields back to datetime objects
                datetime_fields = ['at', 'repliedAt']
                for review in reviews:
                    for field in datetime_fields:
                        if field in review and isinstance(review.get(field), str):
                            try:
                                review[field] = datetime.fromisoformat(review[field])
                            except (ValueError, TypeError):
                                # Keep as string if conversion fails
                                pass

                print(f"  ✓ Loaded {len(reviews)} cached reviews for {app_id}")
                print(f"    Last updated: {last_update.date()}")
                return reviews, last_update
        except Exception as e:
            print(f"  Warning: Could not load cache: {e}")
            return [], None

    return [], None


def save_cached_reviews(app_id: str, reviews: list) -> None:
    """Save reviews to cache for an app."""
    cache_file = get_cache_file_path(app_id)

    try:
        # Convert datetime objects to ISO strings for JSON serialization
        serializable_reviews = []
        for review in reviews:
            review_copy = review.copy()
            # Convert all datetime fields to ISO strings
            for key, value in review_copy.items():
                if isinstance(value, datetime):
                    review_copy[key] = value.isoformat()
            serializable_reviews.append(review_copy)

        cache_data = {
            'app_id': app_id,
            'last_update': datetime.now().isoformat(),
            'total_reviews': len(reviews),
            'reviews': serializable_reviews
        }

        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, indent=2, ensure_ascii=False)

        # Get file size in MB
        file_size_mb = cache_file.stat().st_size / (1024 * 1024)
        print(f"  ✓ Cached {len(reviews)} reviews for {app_id} ({file_size_mb:.2f} MB)")
    except Exception as e:
        print(f"  Warning: Could not save cache: {e}")


def scrape_reviews(app_id: str, start_date: datetime, end_date: datetime) -> Dict[str, List[dict]]:
    """
    Scrape reviews with smart stopping + intelligent caching.
    - Uses cache to avoid re-fetching
    - Stops fetching once past start_date (since sorted NEWEST first)
    - Saves new data to cache for future use
    Returns reviews organized by date.
    """
    print(f"Scraping reviews from {start_date.date()} to {end_date.date()}...")
    reviews_by_date = {}

    print("\nChecking cache...")
    # Load cached reviews
    cached_reviews, last_update = load_cached_reviews(app_id)

    # Combine cached reviews with any new ones
    all_reviews = cached_reviews.copy() if cached_reviews else []
    fetch_from_newest = True

    if cached_reviews:
        # Check if cache is still fresh (less than 1 day old)
        if last_update and (datetime.now() - last_update).days < 1:
            print(f"  Cache is fresh (less than 1 day old)")
            fetch_from_newest = False
        else:
            print(f"  Cache is older than 1 day, updating...")
            fetch_from_newest = True

    if fetch_from_newest:
        print("\nFetching new reviews from Play Store...")
        print("(Smart stopping: will stop once past start_date)\n")

        continuation_token = None
        batch_num = 0
        reached_start_date = False

        try:
            # Fetch reviews until we go past the start_date
            while not reached_start_date:
                batch_num += 1
                print(f"  Batch {batch_num}...", end=" ", flush=True)

                try:
                    # Fetch reviews with pagination
                    # Try US first (more reviews), then IN if that fails
                    countries_to_try = ['us', 'in']
                    result = None
                    last_error = None

                    for country in countries_to_try:
                        try:
                            if batch_num == 1:
                                # First batch without token
                                result, continuation_token = reviews(
                                    app_id,
                                    sort=Sort.NEWEST,
                                    count=500,
                                    lang='en',
                                    country=country
                                )
                            else:
                                # Subsequent batches with continuation token
                                if not continuation_token:
                                    print("(No more reviews)")
                                    break

                                result, continuation_token = reviews(
                                    app_id,
                                    sort=Sort.NEWEST,
                                    count=500,
                                    lang='en',
                                    country=country,
                                    continuation_token=continuation_token
                                )

                            # Success - break out of country loop
                            if result or batch_num > 1:
                                break

                        except Exception as country_error:
                            last_error = country_error
                            if batch_num == 1 and country == countries_to_try[0]:
                                # Try next country
                                continue
                            else:
                                raise country_error

                    # If still no result after trying all countries
                    if not result and batch_num == 1:
                        raise Exception(f"App not available in any region: {last_error}")

                    if not result:
                        print("(Empty batch - stopping)")
                        break

                    batch_fetched = len(result)
                    print(f"{batch_fetched} reviews (Total: {len(all_reviews) + batch_fetched})")

                    # Check if we've gone past the start_date
                    for review in result:
                        review_date = review['at'].date()
                        if review_date < start_date.date():
                            # We've passed the start_date, can stop fetching
                            reached_start_date = True
                            all_reviews.append(review)
                            break
                        all_reviews.append(review)

                    if reached_start_date:
                        print("  ✓ Reached start_date, stopping fetch (smart stopping)")

                except Exception as batch_error:
                    print(f"(Error: {batch_error})")
                    break

            print(f"\n✓ Total reviews fetched: {len(all_reviews)}")

            # Save to cache (only if we actually got reviews)
            if all_reviews:
                save_cached_reviews(app_id, all_reviews)
            else:
                print("  Warning: No reviews fetched, not updating cache")

        except Exception as e:
            print(f"Warning: Scraping error: {e}")
            if not all_reviews:
                print("Proceeding with mock data for demonstration...")
                return generate_mock_reviews(start_date, end_date)

    # Filter by date range
    print(f"\nFiltering reviews for {start_date.date()} to {end_date.date()}...\n")

    for review in all_reviews:
        review_date = review['at'].date()
        if start_date.date() <= review_date <= end_date.date():
            date_str = review_date.isoformat()
            if date_str not in reviews_by_date:
                reviews_by_date[date_str] = []
            reviews_by_date[date_str].append(review)

    reviews_in_range = sum(len(v) for v in reviews_by_date.values())
    print(f"✓ Found {reviews_in_range} reviews within date range")

    # Duplicate detection (optional, controlled by env var)
    enable_dedup = os.getenv('ENABLE_DEDUP', 'false').lower() == 'true'
    if enable_dedup and reviews_in_range > 0:
        try:
            print("\nDuplicate Detection:")
            from utils.duplicate_detector import DuplicateDetector
            from config.embedding_service import EmbeddingService
            from config.cache_db import EmbeddingCache

            # Initialize services with app-specific caching
            embedding_cache = EmbeddingCache()
            embedder = EmbeddingService(use_metal=True, cache=embedding_cache, app_id=app_id)
            threshold = float(os.getenv('DUPLICATE_THRESHOLD', '0.85'))
            detector = DuplicateDetector(embedder, threshold=threshold)

            # Flatten reviews for deduplication
            all_reviews_flat = []
            for date_str, date_reviews in reviews_by_date.items():
                all_reviews_flat.extend(date_reviews)

            # Detect duplicates
            unique, duplicates = detector.detect_duplicates(all_reviews_flat)

            # Rebuild reviews_by_date with only unique reviews
            reviews_by_date_deduped = {}
            unique_ids = {r['reviewId'] for r in unique}

            for date_str, date_reviews in reviews_by_date.items():
                reviews_by_date_deduped[date_str] = [
                    r for r in date_reviews if r['reviewId'] in unique_ids
                ]

            reviews_by_date = reviews_by_date_deduped
            reviews_in_range = sum(len(v) for v in reviews_by_date.values())

            print(f"✓ After deduplication: {reviews_in_range} unique reviews\n")

        except Exception as e:
            print(f"  Warning: Duplicate detection failed ({e}), continuing without dedup...")

    # If no reviews in the specified date range, raise an error
    if reviews_in_range == 0:
        raise ValueError(
            f"No reviews found for {app_id} between {start_date.date()} and {end_date.date()}. "
            "This could mean:\n"
            "  1. The app has very few reviews\n"
            "  2. No reviews exist for this date range\n"
            "  3. The app ID is incorrect\n"
            "  4. Google Play Store rate limiting\n\n"
            "Please try:\n"
            "  - A longer date range (e.g., 90 days)\n"
            "  - A more popular app with more reviews\n"
            "  - Verifying the app ID is correct"
        )

    return reviews_by_date


def generate_mock_reviews(start_date: datetime, end_date: datetime) -> Dict[str, List[dict]]:
    """Generate mock reviews for demonstration purposes."""
    mock_reviews_text = [
        "Delivery was very late, food was cold",
        "Delivery guy was very rude to me",
        "App keeps crashing when I try to checkout",
        "Please bring back the 10 minute delivery option",
        "Food quality has really declined lately",
        "Maps integration not working properly",
        "Great service today, delivery was fast!",
        "Delivery partner was impolite and disrespectful",
        "Food arrived stale, very disappointed",
        "Why did you remove the express delivery?",
        "App is slow and buggy",
        "Delivery person was helpful and friendly",
        "Cold food delivered again",
        "Instamart should be open 24/7",
        "Delivery delay of 2 hours",
    ]

    reviews_by_date = {}
    current_date = start_date
    review_idx = 0

    while current_date <= end_date:
        date_str = current_date.date().isoformat()
        date_reviews = []

        # Add 10-20 reviews per day
        for i in range(10 + (hash(date_str) % 10)):
            date_reviews.append({
                'reviewId': f"review_{date_str}_{i}",
                'userName': f"User{i}",
                'userImage': "",
                'content': mock_reviews_text[(review_idx + i) % len(mock_reviews_text)],
                'score': 1 + (hash(date_str + str(i)) % 5),
                'thumbsUpCount': 0,
                'reviewCreatedVersion': None,
                'at': current_date.replace(hour=10 + (i % 12), minute=i % 60),
                'replyContent': None,
                'repliedAt': None
            })
            review_idx += 1

        reviews_by_date[date_str] = date_reviews
        current_date += timedelta(days=1)

    print(f"✓ Generated {sum(len(v) for v in reviews_by_date.values())} mock reviews")
    return reviews_by_date


def extract_topics_from_review(review_text: str) -> List[dict]:
    """Extract topics from a single review using LLM (Ollama or fallback)."""
    try:
        # Switch to extraction model if using Ollama
        if isinstance(client, OllamaClient):
            client.set_extraction_mode()

        prompt = EXTRACTION_PROMPT.format(review_text=review_text)
        response_text = client.chat(prompt, max_tokens=500, temperature=0.1)
        topics = client.extract_json(response_text)

        # Ensure topics is a list
        if not isinstance(topics, list):
            return []
        return topics

    except Exception as e:
        # Fallback: Use heuristic topic extraction
        return extract_topics_heuristic(review_text)


async def extract_topics_batch_async(reviews_batch: List[dict], async_client: AsyncOllamaClient) -> Dict[int, List[dict]]:
    """Async version: Extract topics from a batch of reviews using a single LLM call."""
    try:
        # Switch to extraction model
        async_client.set_extraction_mode()

        # Build batch string (optimized for larger batches)
        batch_str = ""
        for idx, review in enumerate(reviews_batch):
            content = review.get('content', '')
            if content and len(content) >= 5:
                # Truncate very long reviews to save tokens
                truncated_content = content[:500] if len(content) > 500 else content
                batch_str += f"\n{idx}. {truncated_content}\n"

        if not batch_str:
            return {}

        prompt = BATCH_EXTRACTION_PROMPT.format(reviews_batch=batch_str)
        response_text = await async_client.chat_async(prompt, max_tokens=6000, temperature=0.1)
        result = async_client.extract_json(response_text)

        # Parse results
        topics_by_idx = {}
        if 'reviews' in result:
            for review_result in result['reviews']:
                review_id = int(review_result.get('review_id', -1))
                topics = review_result.get('topics', [])
                if review_id >= 0 and isinstance(topics, list):
                    topics_by_idx[review_id] = topics

        return topics_by_idx

    except Exception as e:
        # Fallback: Extract individually using heuristics
        result = {}
        for idx, review in enumerate(reviews_batch):
            content = review.get('content', '')
            if content and len(content) >= 5:
                result[idx] = extract_topics_heuristic(content)
        return result


def extract_topics_batch(reviews_batch: List[dict]) -> Dict[int, List[dict]]:
    """Extract topics from a batch of reviews using a single LLM call."""
    try:
        # Switch to extraction model if using Ollama
        if isinstance(client, OllamaClient):
            client.set_extraction_mode()

        # Build batch string (optimized for larger batches)
        batch_str = ""
        for idx, review in enumerate(reviews_batch):
            content = review.get('content', '')
            if content and len(content) >= 5:
                # Truncate very long reviews to save tokens
                truncated_content = content[:500] if len(content) > 500 else content
                batch_str += f"\n{idx}. {truncated_content}\n"

        if not batch_str:
            return {}

        prompt = BATCH_EXTRACTION_PROMPT.format(reviews_batch=batch_str)
        response_text = client.chat(prompt, max_tokens=6000, temperature=0.1)
        result = client.extract_json(response_text)

        # Parse results
        topics_by_idx = {}
        if 'reviews' in result:
            for review_result in result['reviews']:
                review_id = int(review_result.get('review_id', -1))
                topics = review_result.get('topics', [])
                if review_id >= 0 and isinstance(topics, list):
                    topics_by_idx[review_id] = topics

        return topics_by_idx

    except Exception as e:
        # Fallback: Extract individually using heuristics
        result = {}
        for idx, review in enumerate(reviews_batch):
            content = review.get('content', '')
            if content and len(content) >= 5:
                result[idx] = extract_topics_heuristic(content)
        return result


def extract_topics_heuristic(review_text: str) -> List[dict]:
    """Fallback: Extract topics using heuristics (for demo without API key)."""
    topics = []
    text_lower = review_text.lower()

    # Define keyword patterns
    patterns = [
        ("rude", "Rude/Unprofessional Delivery Partner", "issue"),
        ("late", "Late Delivery", "issue"),
        ("cold", "Food Delivered Cold", "issue"),
        ("stale", "Food Quality Issue - Stale", "issue"),
        ("crash", "App Crashes/Bugs", "issue"),
        ("slow", "App Performance Issue", "issue"),
        ("10 minute", "Fast Delivery Feature Request", "request"),
        ("24/7", "24/7 Service Request", "request"),
        ("delivery", "Delivery Issue", "issue"),
        ("app", "App Issue", "issue"),
        ("great", "Positive Feedback", "praise"),
        ("good", "Positive Feedback", "praise"),
        ("helpful", "Positive Feedback - Helpful Service", "praise"),
        ("fast", "Fast Delivery Feedback", "praise"),
    ]

    found_topics = set()
    for keyword, topic_name, category in patterns:
        if keyword in text_lower:
            found_topics.add(topic_name)

    # Convert to required format, limit to 5
    for topic_name in list(found_topics)[:5]:
        # Find category for this topic
        category = next((c for k, t, c in patterns if t == topic_name), "feedback")
        topics.append({
            "topic": topic_name,
            "category": category
        })

    return topics if topics else [{"topic": "General Review", "category": "feedback"}]


async def process_batch_wrapper_async(args, async_client: AsyncOllamaClient):
    """Async wrapper function for parallel batch processing."""
    batch_id, batch, batch_start = args
    try:
        print(f"    → Processing batch {batch_id} ({len(batch)} reviews)...", flush=True)
        batch_topics = await extract_topics_batch_async(batch, async_client)
        # Collect topics from this batch
        topics = []
        for idx, review in enumerate(batch):
            if idx in batch_topics:
                extracted = batch_topics[idx]
                for item in extracted:
                    if isinstance(item, dict) and 'topic' in item:
                        topics.append(item['topic'])
        print(f"    ✓ Batch {batch_id} complete ({len(topics)} topics extracted)", flush=True)
        return batch_id, topics, len(batch)
    except Exception as e:
        print(f"    ✗ Batch {batch_id} failed: {e}", flush=True)
        return batch_id, [], len(batch)


async def process_all_batches_async(batches: List[tuple], async_client: AsyncOllamaClient, max_concurrent: int = 16):
    """Process all batches with async concurrency control"""
    # Create semaphore to limit concurrent requests
    semaphore = asyncio.Semaphore(max_concurrent)

    async def bounded_process(batch_args):
        async with semaphore:
            return await process_batch_wrapper_async(batch_args, async_client)

    # Process all batches concurrently
    tasks = [bounded_process(batch_args) for batch_args in batches]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    # Filter out exceptions
    valid_results = []
    for result in results:
        if isinstance(result, Exception):
            print(f"    ✗ Batch failed with exception: {result}")
        else:
            valid_results.append(result)

    return valid_results


def process_batch_wrapper(args):
    """Wrapper function for parallel batch processing."""
    batch_id, batch, batch_start = args
    try:
        print(f"    → Processing batch {batch_id} ({len(batch)} reviews)...", flush=True)
        batch_topics = extract_topics_batch(batch)
        # Collect topics from this batch
        topics = []
        for idx, review in enumerate(batch):
            if idx in batch_topics:
                extracted = batch_topics[idx]
                for item in extracted:
                    if isinstance(item, dict) and 'topic' in item:
                        topics.append(item['topic'])
        print(f"    ✓ Batch {batch_id} complete ({len(topics)} topics extracted)", flush=True)
        return batch_id, topics, len(batch)
    except Exception as e:
        print(f"    ✗ Batch {batch_id} failed: {e}", flush=True)
        return batch_id, [], len(batch)


def extract_all_topics(reviews_by_date: Dict[str, List[dict]], progress_callback=None, job_id=None) -> Dict[str, List[str]]:
    """
    Extract topics from all reviews using ASYNC parallel batch processing with request pipelining.

    Args:
        reviews_by_date: Mapping of date strings to lists of review dictionaries
        progress_callback: Optional function(processed, total) to report progress
        job_id: Optional job ID for cancellation support

    Returns mapping of date -> list of unique topics.
    """
    print("Extracting topics from reviews (using ASYNC parallel batch processing with REQUEST PIPELINING)...")
    topics_by_date = {}
    total_reviews = sum(len(date_reviews) for date_reviews in reviews_by_date.values())

    # AUTO-DETECT optimal settings
    hw_config = get_optimal_concurrency()
    BATCH_SIZE = hw_config['batch_size']
    MAX_CONCURRENT = hw_config['max_concurrent']

    print(f"  Hardware detected:")
    print(f"    CPUs: {hw_config['cpu_count']} cores")
    print(f"    Available RAM: {hw_config['available_memory_gb']} GB")
    print(f"  Optimized configuration:")
    print(f"    Batch size: {BATCH_SIZE} reviews/batch")
    print(f"    Concurrent requests: {MAX_CONCURRENT}")
    print(f"  Total reviews: {total_reviews}")

    # Try to use async client, fall back to sync if not available
    try:
        # Create async client with caching enabled
        enable_cache = os.getenv('ENABLE_CACHE', 'true').lower() == 'true'
        async_client = AsyncOllamaClient(
            base_url=os.getenv('OLLAMA_BASE_URL', 'http://localhost:11434'),
            extraction_model=os.getenv('OLLAMA_EXTRACTION_MODEL', 'qwen2.5:7b'),
            consolidation_model=os.getenv('OLLAMA_CONSOLIDATION_MODEL', 'llama3.1:8b'),
            max_connections=MAX_CONCURRENT,
            enable_cache=enable_cache
        )

        start_time = time.time()
        processed = 0

        # PHASE 3: REQUEST PIPELINING - Build ALL batches across ALL dates upfront
        print("  Building global batch pipeline...")
        all_batches = []
        date_batch_mapping = {}
        global_batch_id = 0

        for date_str in sorted(reviews_by_date.keys()):
            date_reviews = reviews_by_date[date_str]
            date_batch_ids = []

            for batch_start in range(0, len(date_reviews), BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, len(date_reviews))
                batch = date_reviews[batch_start:batch_end]
                all_batches.append((global_batch_id, batch, batch_start))
                date_batch_ids.append(global_batch_id)
                global_batch_id += 1

            date_batch_mapping[date_str] = date_batch_ids

        total_batches = len(all_batches)
        print(f"  Total batches in pipeline: {total_batches}")
        print(f"  Estimated time: ~{int(total_batches / MAX_CONCURRENT * 2)} seconds\n")

        # Process ALL batches in one continuous async pipeline
        async_client.set_extraction_mode()
        all_results = asyncio.run(process_all_batches_async(all_batches, async_client, MAX_CONCURRENT))

        # Organize results back by date
        results_dict = {batch_id: (topics, size) for batch_id, topics, size in all_results}

        for date_str, batch_ids in date_batch_mapping.items():
            date_topics = []
            for batch_id in batch_ids:
                if batch_id in results_dict:
                    topics, batch_size = results_dict[batch_id]
                    date_topics.extend(topics)
                    processed += batch_size

                    # Call progress callback if provided
                    if progress_callback and processed % 50 == 0:
                        progress_callback(processed, total_reviews)

            if date_topics:
                topics_by_date[date_str] = date_topics

        # Cleanup async client
        asyncio.run(async_client.close())

        # Final callback
        if progress_callback:
            progress_callback(total_reviews, total_reviews)

        total_time = time.time() - start_time
        print(f"✓ Extracted topics from {processed} reviews using ASYNC PIPELINED processing")
        print(f"  Total time: {int(total_time)} seconds ({total_time/60:.1f} minutes)")
        print(f"  Average rate: {int(processed/total_time)} reviews/second")

        # Show cache stats if caching was enabled
        if enable_cache and async_client.cache:
            cache_stats = async_client.cache.get_stats()
            print(f"  Cache stats: {cache_stats['total_hits']} hits / {cache_stats['total_entries']} entries")

        return topics_by_date

    except Exception as e:
        print(f"Warning: Async processing failed ({e}), falling back to synchronous processing")
        # Fall back to original synchronous implementation
        return extract_all_topics_sync(reviews_by_date, progress_callback)


def extract_all_topics_sync(reviews_by_date: Dict[str, List[dict]], progress_callback=None) -> Dict[str, List[str]]:
    """Synchronous fallback version of topic extraction"""
    print("Extracting topics from reviews (using synchronous parallel batch processing)...")
    topics_by_date = {}
    total_reviews = sum(len(date_reviews) for date_reviews in reviews_by_date.values())

    BATCH_SIZE = 20
    MAX_WORKERS = 8

    print(f"  Configuration: {BATCH_SIZE} reviews/batch, {MAX_WORKERS} parallel workers")
    print(f"  Total reviews: {total_reviews}")

    start_time = time.time()
    processed = 0

    for date_str in sorted(reviews_by_date.keys()):
        date_reviews = reviews_by_date[date_str]

        # Create batches for this date
        batches = []
        batch_id = 0
        for batch_start in range(0, len(date_reviews), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(date_reviews))
            batch = date_reviews[batch_start:batch_end]
            batches.append((batch_id, batch, batch_start))
            batch_id += 1

        # Process batches in parallel
        date_topics = []
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_batch_wrapper, batch_args): batch_args[0]
                      for batch_args in batches}

            for future in as_completed(futures):
                batch_id, topics, batch_size = future.result()
                date_topics.extend(topics)
                processed += batch_size

                if progress_callback and processed % 50 == 0:
                    progress_callback(processed, total_reviews)

        if date_topics:
            topics_by_date[date_str] = date_topics

    if progress_callback:
        progress_callback(total_reviews, total_reviews)

    total_time = time.time() - start_time
    print(f"✓ Extracted topics from {processed} reviews")
    print(f"  Total time: {int(total_time)} seconds ({total_time/60:.1f} minutes)")
    print(f"  Average rate: {int(processed/total_time)} reviews/second")
    return topics_by_date


def normalize_topic(topic: str) -> str:
    """Normalize topic text for better consolidation matching."""
    # Convert to lowercase
    normalized = topic.lower().strip()

    # Remove extra whitespace
    normalized = re.sub(r'\s+', ' ', normalized)

    # Normalize common word variations
    replacements = {
        r'\b(is|are|was|were|be|being|been)\b': '',  # Remove "to be" verbs
        r'\b(a|an|the)\b': '',  # Remove articles
        r'\b(very|extremely|really|so|too)\b': '',  # Remove intensifiers
        r'\bhas\b': '',
        r'\bcontains\b': '',
        r'\bdelivery partner\b': 'delivery partner',  # Normalize to singular form
        r'\bdelivery guy\b': 'delivery partner',
        r'\bdelivery person\b': 'delivery partner',
        r'\brider\b': 'delivery partner',
        r'\b(\d+)\s*hour[s]?\b': r'\1 hour',  # Normalize hours
        r'\b(\d+)\s*min(ute)?[s]?\b': r'\1 minute',  # Normalize minutes
    }

    for pattern, replacement in replacements.items():
        normalized = re.sub(pattern, replacement, normalized)

    # Remove extra spaces created by replacements
    normalized = re.sub(r'\s+', ' ', normalized).strip()

    return normalized


def consolidate_topics(all_topics: List[str], app_id: str = None) -> Dict[str, List[str]]:
    """
    Consolidate similar topics using HYBRID approach:
    1. Try embedding-based clustering (fast, 240x speedup) - default
    2. Fallback to LLM consolidation if embeddings disabled

    Returns mapping of canonical_topic -> list of variations.
    """
    if not all_topics:
        return {}

    # Check if embedding clustering is enabled
    use_embeddings = os.getenv('ENABLE_EMBEDDING_CLUSTERING', 'true').lower() == 'true'

    if use_embeddings:
        try:
            print(f"Consolidating {len(all_topics)} topics using embedding clustering (fast)...")

            # Initialize services
            from config.embedding_service import EmbeddingService
            from config.cache_db import EmbeddingCache
            from ml.topic_clustering import TopicClusterer

            embedding_cache = EmbeddingCache()
            embedder = EmbeddingService(use_metal=True, cache=embedding_cache, app_id=app_id)
            clusterer = TopicClusterer(embedder, min_cluster_size=3)

            # Cluster topics
            canonical_mapping = clusterer.cluster_topics(all_topics)

            print(f"✓ Consolidated to {len(canonical_mapping)} canonical topics using embeddings")

            return canonical_mapping

        except Exception as e:
            print(f"  Warning: Embedding clustering failed ({e}), falling back to LLM...")
            # Fall through to LLM-based consolidation

    # Fallback: LLM-based consolidation (original implementation)
    return consolidate_topics_llm(all_topics)


def consolidate_topics_llm(all_topics: List[str]) -> Dict[str, List[str]]:
    """
    Original LLM-based consolidation (kept as fallback)
    Consolidate similar topics using LLM with aggressive grouping.
    """
    if not all_topics:
        return {}

    # Remove duplicates but keep count
    unique_topics = list(set(all_topics))
    print(f"Consolidating {len(unique_topics)} unique topics using LLM (slow)...")

    # Pre-process: Group exact matches after normalization
    normalized_groups = {}
    for topic in unique_topics:
        normalized = normalize_topic(topic)
        if normalized not in normalized_groups:
            normalized_groups[normalized] = []
        normalized_groups[normalized].append(topic)

    print(f"  After normalization: {len(normalized_groups)} topic groups")

    # Create topic list string (use one representative from each normalized group)
    representative_topics = [topics[0] for topics in normalized_groups.values()]
    topics_str = "\n".join([f"- {t}" for t in representative_topics[:200]])  # Increased limit
    if len(representative_topics) > 200:
        topics_str += f"\n... and {len(representative_topics) - 200} more topics"

    try:
        # Switch to consolidation model if using Ollama (high quality needed here)
        if isinstance(client, OllamaClient):
            client.set_consolidation_mode()

        prompt = CONSOLIDATION_PROMPT.format(
            num_topics=len(representative_topics),
            topics_list=topics_str
        )

        response_text = client.chat(prompt, max_tokens=4000, temperature=0.1)
        result = client.extract_json(response_text)

        # Build canonical mapping
        canonical_mapping = {}
        if 'canonical_topics' in result:
            for item in result['canonical_topics']:
                canonical = item.get('canonical_name', '')
                variations = item.get('variations', [])

                if canonical:
                    # Add all variations from normalized groups
                    all_variations = []
                    for variation in variations:
                        # Find all topics in the same normalized group
                        norm_var = normalize_topic(variation)
                        if norm_var in normalized_groups:
                            all_variations.extend(normalized_groups[norm_var])

                    canonical_mapping[canonical] = list(set(all_variations))

        print(f"✓ Consolidated to {len(canonical_mapping)} canonical topics")
        return canonical_mapping

    except Exception as e:
        print(f"  Warning: Consolidation error: {e}")
        # Fallback: use heuristic consolidation
        return consolidate_topics_heuristic(unique_topics)


def consolidate_topics_heuristic(unique_topics: List[str]) -> Dict[str, List[str]]:
    """Fallback: Consolidate topics using heuristics with AGGRESSIVE merging."""
    # Define VERY aggressive consolidation rules
    consolidation_rules = {
        "Positive feedback": [
            "positive", "good", "great", "excellent", "amazing", "awesome", "love",
            "best", "helpful", "friendly", "fast", "quick", "perfect", "satisfied"
        ],
        "Delivery partner unprofessional": [
            "rude", "impolite", "unprofessional", "disrespectful", "behavior", "attitude"
        ],
        "Delivery delay": [
            "late", "delay", "slow", "hour", "wait", "time", "delayed"
        ],
        "Food temperature issues": [
            "cold", "hot", "warm", "temperature", "lukewarm"
        ],
        "Food freshness issues": [
            "stale", "spoiled", "rotten", "old", "fresh", "bad quality"
        ],
        "App crashes/freezes": [
            "crash", "freeze", "stuck", "not working", "not responding", "hang"
        ],
        "App performance issues": [
            "slow", "lag", "bug", "glitch", "issue", "problem", "error"
        ],
        "10 minute delivery removed": [
            "10 minute", "bolt", "express", "fast delivery"
        ],
        "24/7 service request": [
            "24/7", "24 hour", "all night", "late night", "instamart"
        ],
        "Missing items": [
            "missing", "forgot", "didn't receive", "not delivered"
        ],
        "Wrong order": [
            "wrong", "incorrect", "mistake", "different"
        ],
        "Payment issues": [
            "payment", "charge", "refund", "money", "price", "cost"
        ],
    }

    # Map topics to canonical
    canonical_mapping = {}
    mapped_topics = set()

    for canonical, keywords in consolidation_rules.items():
        matching_topics = []
        for topic in unique_topics:
            topic_lower = topic.lower()
            # Check if any keyword appears in the topic
            if any(keyword in topic_lower for keyword in keywords):
                matching_topics.append(topic)
                mapped_topics.add(topic)
        
        if matching_topics:
            canonical_mapping[canonical] = matching_topics

    # Add unmapped topics as their own canonical topics
    for topic in unique_topics:
        if topic not in mapped_topics:
            canonical_mapping[topic] = [topic]

    return canonical_mapping


def map_topics_to_canonical(topics_by_date: Dict[str, List[str]],
                           canonical_mapping: Dict[str, List[str]],
                           app_id: str = None) -> tuple[Dict[str, Dict[str, int]], Dict[str, str]]:
    """
    Map extracted topics to canonical topics using HYBRID approach:
    1. Try embedding similarity (accurate, fixes duplicate positive feedback issue)
    2. Fallback to fuzzy matching if embeddings disabled

    Returns: (date -> canonical_topic -> count mapping, unmapped_topics -> suggested_canonical mapping)
    """
    use_embeddings = os.getenv('ENABLE_EMBEDDING_CLUSTERING', 'true').lower() == 'true'

    if use_embeddings:
        try:
            from config.embedding_service import EmbeddingService
            from config.cache_db import EmbeddingCache

            embedding_cache = EmbeddingCache()
            embedder = EmbeddingService(use_metal=True, cache=embedding_cache, app_id=app_id)

            return map_topics_to_canonical_embedding(
                topics_by_date,
                canonical_mapping,
                embedder,
                similarity_threshold=float(os.getenv('TOPIC_SIMILARITY_THRESHOLD', '0.70'))
            )
        except Exception as e:
            print(f"  Warning: Embedding-based mapping failed ({e}), using fuzzy matching...")
            # Fall through to original implementation

    # Fallback: Original fuzzy matching
    return map_topics_to_canonical_fuzzy(topics_by_date, canonical_mapping)


def map_topics_to_canonical_embedding(
    topics_by_date: Dict[str, List[str]],
    canonical_mapping: Dict[str, List[str]],
    embedding_service,
    similarity_threshold: float = 0.70
) -> tuple[Dict[str, Dict[str, int]], Dict[str, str]]:
    """
    NEW: Embedding-based topic mapping (much better than substring matching)
    Fixes the duplicate "positive feedback" issue
    """
    # Build list of canonical topics
    canonical_topics = list(canonical_mapping.keys())

    # Generate embeddings for canonical topics (cached)
    print(f"  Generating embeddings for {len(canonical_topics)} canonical topics...")
    canonical_embeddings = embedding_service.encode(canonical_topics, batch_size=128)

    # Track unmapped topics
    unmapped_topics = {}

    # Count topics by date
    result = {}

    for date_str in sorted(topics_by_date.keys()):
        result[date_str] = {}
        date_topics = topics_by_date[date_str]

        if not date_topics:
            continue

        # Generate embeddings for this date's topics
        topic_embeddings = embedding_service.encode(date_topics, batch_size=128)

        # Compute similarity matrix (topics × canonicals)
        from sklearn.metrics.pairwise import cosine_similarity
        similarity_matrix = cosine_similarity(topic_embeddings, canonical_embeddings)

        # Map each topic to best canonical match
        for topic_idx, topic in enumerate(date_topics):
            similarities = similarity_matrix[topic_idx]
            best_idx = similarities.argmax()
            best_similarity = similarities[best_idx]

            if best_similarity >= similarity_threshold:
                canonical = canonical_topics[best_idx]
            else:
                # No good match - use original topic
                canonical = topic
                if topic not in unmapped_topics:
                    # Suggest best match even if below threshold
                    unmapped_topics[topic] = canonical_topics[best_idx]

            result[date_str][canonical] = result[date_str].get(canonical, 0) + 1

    if unmapped_topics:
        print(f"  ⚠ {len(unmapped_topics)} topics below similarity threshold ({similarity_threshold})")

    return result, unmapped_topics


def map_topics_to_canonical_fuzzy(topics_by_date: Dict[str, List[str]],
                                   canonical_mapping: Dict[str, List[str]]) -> tuple[Dict[str, Dict[str, int]], Dict[str, str]]:
    """
    FALLBACK: Original fuzzy matching (kept for backward compatibility)
    """
    # Build reverse mapping: variation -> canonical (case-insensitive)
    variation_to_canonical = {}
    for canonical, variations in canonical_mapping.items():
        for variation in variations:
            variation_to_canonical[variation.lower()] = canonical

    # Track unmapped topics
    unmapped_topics = {}

    # Count topics by date
    result = {}
    for date_str in sorted(topics_by_date.keys()):
        result[date_str] = {}
        for topic in topics_by_date[date_str]:
            topic_lower = topic.lower()

            # Try exact match first
            if topic_lower in variation_to_canonical:
                canonical = variation_to_canonical[topic_lower]
            else:
                # Try fuzzy matching - check if topic contains any canonical variation
                canonical = None
                for var, can in variation_to_canonical.items():
                    if var in topic_lower or topic_lower in var:
                        canonical = can
                        break

                # If still not found, use original topic but track it
                if canonical is None:
                    canonical = topic
                    if topic not in unmapped_topics:
                        # Try to find best canonical match based on keywords
                        best_match = find_best_canonical_match(topic, canonical_mapping)
                        unmapped_topics[topic] = best_match

            result[date_str][canonical] = result[date_str].get(canonical, 0) + 1

    return result, unmapped_topics


def find_best_canonical_match(topic: str, canonical_mapping: Dict[str, List[str]]) -> str:
    """Find the best canonical topic match for an unmapped topic using keyword similarity."""
    topic_lower = topic.lower()
    topic_words = set(re.findall(r'\w+', topic_lower))
    
    best_match = topic  # Default to original
    best_score = 0
    
    for canonical, variations in canonical_mapping.items():
        # Get all words from all variations
        all_words = set()
        for var in variations:
            all_words.update(re.findall(r'\w+', var.lower()))
        
        # Calculate word overlap
        overlap = len(topic_words & all_words)
        if overlap > best_score:
            best_score = overlap
            best_match = canonical
    
    return best_match if best_score > 0 else topic


def extract_app_id_from_link(link: str) -> str:
    """Extract app package ID from a Play Store link or return as-is if it's already a package ID."""
    if not link:
        return None

    link = link.strip()

    # Check if it's a Play Store URL
    if 'play.google.com' in link:
        # Extract package ID from URL like: https://play.google.com/store/apps/details?id=com.example.app
        if 'id=' in link:
            app_id = link.split('id=')[1].split('&')[0]
            return app_id

    # Otherwise assume it's already a package ID
    return link


def generate_trend_report(canonical_counts: Dict[str, Dict[str, int]],
                         target_date: datetime,
                         output_file: str,
                         canonical_mapping: Dict[str, List[str]],
                         unmapped_topics: Dict[str, str]) -> None:
    """Generate Excel trend report with Topic x Date matrix and validation."""
    print(f"Generating trend report for {target_date.date()}...")

    # Ensure we have 30 days of data
    start_date = target_date - timedelta(days=29)
    date_range = [start_date + timedelta(days=i) for i in range(30)]
    date_strs = [d.strftime("%Y-%m-%d") for d in date_range]

    # Collect all topics that appear in the data
    all_topics_in_data = set()
    for date_str in date_strs:
        if date_str in canonical_counts:
            all_topics_in_data.update(canonical_counts[date_str].keys())

    # VALIDATION: Check for discrepancies
    print(f"\n📊 VALIDATION REPORT:")
    print(f"  Expected canonical topics: {len(canonical_mapping)}")
    print(f"  Topics appearing in Excel: {len(all_topics_in_data)}")
    
    if len(all_topics_in_data) != len(canonical_mapping):
        print(f"  ⚠️  WARNING: Mismatch detected ({len(all_topics_in_data)} vs {len(canonical_mapping)})")
        
        # Find topics in data but not in canonical mapping
        extra_topics = all_topics_in_data - set(canonical_mapping.keys())
        if extra_topics:
            print(f"\n  ❌ Topics in Excel but NOT in canonical mapping ({len(extra_topics)}):")
            for topic in sorted(extra_topics)[:10]:  # Show first 10
                suggestion = unmapped_topics.get(topic, "Unknown")
                print(f"    - '{topic}' (suggested: '{suggestion}')")
            if len(extra_topics) > 10:
                print(f"    ... and {len(extra_topics) - 10} more")
        
        # Find canonical topics that never appear
        missing_topics = set(canonical_mapping.keys()) - all_topics_in_data
        if missing_topics:
            print(f"\n  ℹ️  Canonical topics that never appeared ({len(missing_topics)}):")
            for topic in sorted(missing_topics)[:5]:
                print(f"    - '{topic}'")
            if len(missing_topics) > 5:
                print(f"    ... and {len(missing_topics) - 5} more")
    else:
        print(f"  ✅ Perfect match! All topics mapped correctly.")

    # Count topics with only 1 occurrence total
    single_occurrence_topics = []
    for topic in all_topics_in_data:
        total_count = sum(canonical_counts.get(date_str, {}).get(topic, 0) for date_str in date_strs)
        if total_count == 1:
            single_occurrence_topics.append(topic)
    
    if single_occurrence_topics:
        print(f"\n  ⚠️  Topics with only 1 occurrence (likely unmapped) ({len(single_occurrence_topics)}):")
        for topic in sorted(single_occurrence_topics)[:10]:
            suggestion = unmapped_topics.get(topic, "Unknown")
            print(f"    - '{topic}' (suggested: '{suggestion}')")
        if len(single_occurrence_topics) > 10:
            print(f"    ... and {len(single_occurrence_topics) - 10} more")

    print()  # Blank line before Excel generation

    # Sort topics: canonical first, then unmapped
    canonical_topics = sorted([t for t in all_topics_in_data if t in canonical_mapping])
    unmapped_in_data = sorted([t for t in all_topics_in_data if t not in canonical_mapping])
    all_topics = canonical_topics + unmapped_in_data

    # Create matrix
    data = []
    for topic in all_topics:
        row = [topic]
        for date_str in date_strs:
            count = canonical_counts.get(date_str, {}).get(topic, 0)
            row.append(count)
        data.append(row)

    # Create DataFrame
    columns = ["Topic"] + [d.strftime("%b %d") for d in date_range]
    df = pd.DataFrame(data, columns=columns)

    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Trend Report"

    # Add headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data with alternating colors and highlight unmapped topics
    light_fill = PatternFill(start_color="E8EFF7", end_color="E8EFF7", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")  # Yellow for unmapped
    
    for row_idx, row_data in enumerate(data, 2):
        topic_name = row_data[0]
        is_unmapped = topic_name not in canonical_mapping
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Highlight unmapped topics
            if is_unmapped and col_idx == 1:
                cell.fill = warning_fill
                cell.font = Font(italic=True)
            elif row_idx % 2 == 0:
                cell.fill = light_fill
            
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions['A'].width = 40
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for col_idx in range(1, len(columns)):
        if col_idx < len(col_letters):
            ws.column_dimensions[col_letters[col_idx]].width = 12

    # Add legend for unmapped topics
    if unmapped_in_data:
        legend_row = len(data) + 3
        ws.cell(row=legend_row, column=1, value="⚠️ Yellow-highlighted topics are unmapped (not in canonical list)")
        ws.cell(row=legend_row, column=1).font = Font(italic=True, color="FF6B35")

    # Save
    wb.save(output_file)
    print(f"✓ Report saved to {output_file}")
    
    if unmapped_in_data:
        print(f"  Note: {len(unmapped_in_data)} unmapped topics highlighted in yellow")


def main():
    """Main execution flow."""
    parser = argparse.ArgumentParser(description="App Store Review Trend Analysis")
    parser.add_argument("--target-date", type=str, default=None,
                       help="Target date (YYYY-MM-DD). Default: today")
    parser.add_argument("--days", type=int, default=30,
                       help="Number of days to analyze (default: 30)")
    parser.add_argument("--app-id", type=str, default=None,
                       help="App package ID or Play Store link")

    args = parser.parse_args()

    print("=" * 60)
    print("App Store Review Trend Analysis")
    print("=" * 60)

    # Prompt for app ID if not provided via command line
    if args.app_id:
        app_id = extract_app_id_from_link(args.app_id)
    else:
        while True:
            app_input = input("\nEnter app package ID (e.g., in.swiggy.android) or Play Store link,\nor press Enter for Swiggy: ").strip()

            if not app_input:
                app_id = SWIGGY_APP_ID
                print(f"Using default app: Swiggy ({app_id})")
                break
            else:
                app_id = extract_app_id_from_link(app_input)
                if app_id:
                    print(f"Using app: {app_id}")
                    break
                else:
                    print("Invalid input. Please provide a valid package ID or Play Store link.")

    # Prompt for target date if not provided via command line
    if args.target_date:
        target_date = datetime.strptime(args.target_date, "%Y-%m-%d")
    else:
        while True:
            try:
                date_input = input("\nEnter target date (YYYY-MM-DD) or press Enter for today: ").strip()

                if not date_input:
                    target_date = datetime.now()
                    print(f"Using today's date: {target_date.date()}")
                else:
                    target_date = datetime.strptime(date_input, "%Y-%m-%d")
                    print(f"Using target date: {target_date.date()}")
                break
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD format.")

    # Calculate date range (30 days including target date)
    end_date = target_date
    start_date = target_date - timedelta(days=args.days - 1)

    print(f"Analysis Period: {start_date.date()} to {end_date.date()}")
    print()

    # Phase 1: Data Collection
    print("PHASE 1: Data Collection")
    print("-" * 40)
    reviews_by_date = scrape_reviews(app_id, start_date, end_date)
    print()

    # Phase 2: Topic Extraction
    print("PHASE 2: Topic Extraction")
    print("-" * 40)
    topics_by_date = extract_all_topics(reviews_by_date)
    print()

    # Phase 3: Topic Consolidation
    print("PHASE 3: Topic Consolidation")
    print("-" * 40)
    all_extracted_topics = []
    for topics in topics_by_date.values():
        all_extracted_topics.extend(topics)

    canonical_mapping = consolidate_topics(all_extracted_topics, app_id=app_id)
    
    # Debug: Print canonical topics
    print(f"\n📋 Canonical Topics ({len(canonical_mapping)}):")
    for idx, canonical in enumerate(sorted(canonical_mapping.keys()), 1):
        variation_count = len(canonical_mapping[canonical])
        print(f"  {idx}. {canonical} ({variation_count} variations)")
    print()

    # Phase 4: Generate Trend Matrix
    print("PHASE 4: Trend Analysis")
    print("-" * 40)
    canonical_counts, unmapped_topics = map_topics_to_canonical(topics_by_date, canonical_mapping, app_id=app_id)
    
    # Debug: Print unmapped topics if any
    if unmapped_topics:
        print(f"\n⚠️  Found {len(unmapped_topics)} unmapped topics:")
        for topic, suggestion in list(unmapped_topics.items())[:10]:
            print(f"  - '{topic}' → suggested: '{suggestion}'")
        if len(unmapped_topics) > 10:
            print(f"  ... and {len(unmapped_topics) - 10} more")
    print()

    # Phase 5: Generate Report
    print("PHASE 5: Report Generation")
    print("-" * 40)
    # Extract app name from app_id (e.g., "in.swiggy.android" → "swiggy")
    app_name = app_id.split('.')[-2] if '.' in app_id else app_id
    output_file = OUTPUT_DIR / f"{app_name}_trend_report_{target_date.strftime('%Y-%m-%d')}.xlsx"
    generate_trend_report(canonical_counts, target_date, str(output_file), canonical_mapping, unmapped_topics)

    print()
    print("=" * 60)
    print("✓ Analysis Complete!")
    print("=" * 60)
    print(f"Output saved to: {output_file}")


if __name__ == "__main__":
    main()